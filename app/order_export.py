from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Iterable, Mapping
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
VIETNAM_TIMEZONE = ZoneInfo("Asia/Ho_Chi_Minh")

STATUS_LABELS = {
    "draft": "Cần kiểm tra (dữ liệu cũ)",
    "missing_info": "Thiếu thông tin",
    "ready_to_confirm": "Chờ khách xác nhận",
    "confirmed": "Chờ đóng gói",
    "packed": "Đã đóng gói",
    "shipping": "Đang giao hàng",
    "completed": "Hoàn thành",
    "cancelled": "Đã hủy",
}
STATUS_COLORS = {
    "draft": "FFF3CD",
    "missing_info": "FDE68A",
    "ready_to_confirm": "DBEAFE",
    "confirmed": "DCFCE7",
    "packed": "CFFAFE",
    "shipping": "FEF3C7",
    "completed": "D1FAE5",
    "cancelled": "FEE2E2",
}
MISSING_FIELD_LABELS = {
    "customer_phone": "Thiếu SĐT",
    "shipping_address": "Thiếu địa chỉ",
    "product_name": "Thiếu sản phẩm",
    "quantity": "Thiếu số lượng",
    "total_price": "Thiếu tổng tiền",
}


def build_packing_workbook(
    orders: Iterable[Mapping[str, Any]],
    *,
    filter_label: str = "Tất cả đơn hàng",
    generated_at: datetime | None = None,
) -> bytes:
    rows = list(orders)
    generated = _local_datetime(generated_at or datetime.now(timezone.utc))
    workbook = Workbook()
    workbook.properties.creator = "Voice AI Sales"
    workbook.properties.title = "Phiếu đóng gói đơn hàng"
    workbook.properties.subject = filter_label

    sheet = workbook.active
    sheet.title = "Phiếu đóng gói"
    _build_packing_sheet(sheet, rows, filter_label, generated)

    summary = workbook.create_sheet("Tổng hợp sản phẩm")
    _build_product_summary(summary, rows, generated)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def packing_workbook_filename(generated_at: datetime | None = None) -> str:
    generated = _local_datetime(generated_at or datetime.now(timezone.utc))
    return f"phieu-dong-goi-{generated:%Y%m%d-%H%M}.xlsx"


def _build_packing_sheet(sheet, orders, filter_label: str, generated: datetime) -> None:
    green = "166534"
    dark_green = "14532D"
    pale_green = "F0FDF4"
    white = "FFFFFF"
    grey = "64748B"
    thin = Side(style="thin", color="D9E2EC")

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A6"
    sheet.sheet_properties.tabColor = green
    sheet.merge_cells("A1:L1")
    sheet["A1"] = "PHIẾU ĐÓNG GÓI ĐƠN HÀNG"
    sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=white)
    sheet["A1"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 38

    sheet.merge_cells("A2:L2")
    sheet["A2"] = (
        f"Bộ lọc: {_excel_text(filter_label)}   •   "
        f"Xuất lúc: {generated:%H:%M %d/%m/%Y} (GMT+7)"
    )
    sheet["A2"].font = Font(name="Aptos", size=10, color=grey, italic=True)
    sheet["A2"].fill = PatternFill("solid", fgColor="F8FAFC")
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24

    total_units = sum(_int(order.get("quantity")) for order in orders)
    total_value = sum(_int(order.get("total_price")) for order in orders)
    confirmed_count = sum(order.get("status") == "confirmed" for order in orders)
    metrics = (
        ("A3:C3", f"TỔNG ĐƠN\n{len(orders):,}"),
        ("D3:F3", f"TỔNG SẢN PHẨM\n{total_units:,}"),
        ("G3:I3", f"TỔNG GIÁ TRỊ\n{total_value:,} MMK"),
        ("J3:L3", f"CHỜ ĐÓNG GÓI\n{confirmed_count:,}"),
    )
    for cell_range, text in metrics:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = text
        cell.font = Font(name="Aptos", size=11, bold=True, color=dark_green)
        cell.fill = PatternFill("solid", fgColor=pale_green)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.row_dimensions[3].height = 44
    sheet.row_dimensions[4].height = 8

    headers = (
        "STT",
        "Mã đơn",
        "Ngày tạo",
        "Khách hàng",
        "Số điện thoại",
        "Sản phẩm / Combo",
        "SL",
        "Đơn giá (MMK)",
        "Tổng tiền (MMK)",
        "Địa chỉ giao hàng",
        "Trạng thái",
        "Thông tin còn thiếu / Ghi chú",
    )
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=column, value=title)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.row_dimensions[5].height = 34

    for row_number, order in enumerate(orders, start=6):
        status = str(order.get("status") or "draft")
        values = (
            row_number - 5,
            f"#{order.get('id', '')}",
            _parse_order_datetime(order.get("created_at")),
            _excel_text(order.get("customer_name") or "Chưa có tên"),
            _excel_text(order.get("customer_phone")),
            _excel_text(order.get("product_name")),
            _int(order.get("quantity")),
            _int(order.get("unit_price")),
            _int(order.get("total_price")),
            _excel_text(order.get("shipping_address")),
            STATUS_LABELS.get(status, _excel_text(status)),
            _missing_fields_text(order.get("missing_fields")),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 2, 3, 5, 7, 8, 9, 11} else "left",
                vertical="center",
                wrap_text=column in {4, 6, 10, 11, 12},
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row_number % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
        sheet.cell(row=row_number, column=3).number_format = "dd/mm/yyyy hh:mm"
        sheet.cell(row=row_number, column=5).number_format = "@"
        sheet.cell(row=row_number, column=7).number_format = "0"
        sheet.cell(row=row_number, column=8).number_format = '#,##0 "MMK"'
        sheet.cell(row=row_number, column=9).number_format = '#,##0 "MMK"'
        status_cell = sheet.cell(row=row_number, column=11)
        status_cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "E2E8F0"))
        status_cell.font = Font(name="Aptos", size=10, bold=True, color="1F2937")
        sheet.row_dimensions[row_number].height = 42

    last_row = max(6, 5 + len(orders))
    if orders:
        table = Table(displayName="PackingOrders", ref=f"A5:L{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.auto_filter.ref = f"A5:L{last_row}"
    else:
        sheet.merge_cells("A6:L8")
        sheet["A6"] = "Không có đơn hàng phù hợp với bộ lọc hiện tại."
        sheet["A6"].font = Font(name="Aptos", size=12, italic=True, color=grey)
        sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")

    widths = (6, 11, 18, 24, 18, 32, 8, 17, 18, 44, 24, 32)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width

    sheet.print_title_rows = "1:5"
    sheet.print_area = f"A1:L{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "Trang &P / &N"
    sheet.oddFooter.right.text = "Voice AI Sales"
    sheet.oddFooter.center.size = 9
    sheet.oddFooter.right.size = 9
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45


def _build_product_summary(sheet, orders, generated: datetime) -> None:
    green = "166534"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2EC")
    grouped: dict[str, dict[str, int]] = defaultdict(
        lambda: {"orders": 0, "quantity": 0, "value": 0}
    )
    for order in orders:
        name = _excel_text(order.get("product_name") or "Chưa phân loại")
        grouped[name]["orders"] += 1
        grouped[name]["quantity"] += _int(order.get("quantity"))
        grouped[name]["value"] += _int(order.get("total_price"))

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.sheet_properties.tabColor = "22C55E"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "TỔNG HỢP SẢN PHẨM CẦN ĐÓNG GÓI"
    sheet["A1"].font = Font(name="Aptos Display", size=17, bold=True, color=white)
    sheet["A1"].fill = PatternFill("solid", fgColor=green)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:D2")
    sheet["A2"] = f"Cập nhật lúc {generated:%H:%M %d/%m/%Y} (GMT+7)"
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color="64748B")

    headers = ("Sản phẩm / Combo", "Số đơn", "Tổng số lượng", "Tổng giá trị (MMK)")
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=title)
        cell.font = Font(name="Aptos", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_number, name in enumerate(sorted(grouped, key=str.casefold), start=4):
        data = grouped[name]
        for column, value in enumerate(
            (name, data["orders"], data["quantity"], data["value"]), start=1
        ):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "center",
                vertical="center",
                wrap_text=column == 1,
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row_number % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0FDF4")
        sheet.cell(row=row_number, column=4).number_format = '#,##0 "MMK"'

    last_row = max(3, len(grouped) + 3)
    if grouped:
        table = Table(displayName="PackingProductSummary", ref=f"A3:D{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        sheet.add_table(table)
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 22
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _missing_fields_text(value: Any) -> str:
    fields = value if isinstance(value, list) else str(value or "").split(",")
    labels = [
        MISSING_FIELD_LABELS.get(str(field).strip(), str(field).strip())
        for field in fields
        if str(field).strip()
    ]
    return _excel_text(", ".join(labels))


def _excel_text(value: Any) -> str:
    text = str(value or "").strip()
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _parse_order_datetime(value: Any) -> datetime | str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return _local_datetime(datetime.fromisoformat(text.replace("Z", "+00:00"))).replace(
            tzinfo=None
        )
    except ValueError:
        return _excel_text(text)


def _local_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(VIETNAM_TIMEZONE)


def _int(value: Any) -> int:
    try:
        return max(0, int(value or 0))
    except (TypeError, ValueError):
        return 0


def _column_letter(index: int) -> str:
    return chr(64 + index)
